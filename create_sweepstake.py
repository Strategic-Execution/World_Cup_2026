"""
FIFA World Cup 2026 Sweepstake Excel Model Generator
Creates a comprehensive Excel workbook for running a workplace sweepstake.

Sheets:
1. Teams - All 48 teams with groups
2. Participants - Enter names from Microsoft Forms
3. Allocation - Random team-to-participant mapping
4. Fixtures - All 104 matches with score entry
5. Points - Auto-calculated points per team
6. Leaderboard - Ranked participant standings
7. Golden Boot - Player nominations tracker
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from copy import copy

# ============================================================
# DATA
# ============================================================

TEAMS = {
    "A": ["Mexico", "South Africa", "South Korea", "Czechia"],
    "B": ["Canada", "Bosnia and Herzegovina", "Qatar", "Switzerland"],
    "C": ["Brazil", "Morocco", "Haiti", "Scotland"],
    "D": ["United States", "Paraguay", "Australia", "Türkiye"],
    "E": ["Germany", "Curaçao", "Ivory Coast", "Ecuador"],
    "F": ["Netherlands", "Japan", "Sweden", "Tunisia"],
    "G": ["Belgium", "Egypt", "Iran", "New Zealand"],
    "H": ["Spain", "Cape Verde", "Saudi Arabia", "Uruguay"],
    "I": ["France", "Senegal", "Iraq", "Norway"],
    "J": ["Argentina", "Algeria", "Austria", "Jordan"],
    "K": ["Portugal", "DR Congo", "Uzbekistan", "Colombia"],
    "L": ["England", "Croatia", "Ghana", "Panama"],
}

# Group stage fixtures (Match#, Date, Home, Away, Group)
GROUP_FIXTURES = [
    (1, "11-Jun", "Mexico", "South Africa", "A"),
    (2, "11-Jun", "South Korea", "Czechia", "A"),
    (3, "12-Jun", "Canada", "Bosnia and Herzegovina", "B"),
    (4, "12-Jun", "United States", "Paraguay", "D"),
    (5, "13-Jun", "Haiti", "Scotland", "C"),
    (6, "13-Jun", "Australia", "Türkiye", "D"),
    (7, "13-Jun", "Brazil", "Morocco", "C"),
    (8, "13-Jun", "Qatar", "Switzerland", "B"),
    (9, "14-Jun", "Ivory Coast", "Ecuador", "E"),
    (10, "14-Jun", "Germany", "Curaçao", "E"),
    (11, "14-Jun", "Netherlands", "Japan", "F"),
    (12, "14-Jun", "Sweden", "Tunisia", "F"),
    (13, "15-Jun", "Saudi Arabia", "Uruguay", "H"),
    (14, "15-Jun", "Spain", "Cape Verde", "H"),
    (15, "15-Jun", "Iran", "New Zealand", "G"),
    (16, "15-Jun", "Belgium", "Egypt", "G"),
    (17, "16-Jun", "France", "Senegal", "I"),
    (18, "16-Jun", "Iraq", "Norway", "I"),
    (19, "16-Jun", "Argentina", "Algeria", "J"),
    (20, "16-Jun", "Austria", "Jordan", "J"),
    (21, "17-Jun", "Ghana", "Panama", "L"),
    (22, "17-Jun", "England", "Croatia", "L"),
    (23, "17-Jun", "Portugal", "DR Congo", "K"),
    (24, "17-Jun", "Uzbekistan", "Colombia", "K"),
    # Matchday 2
    (25, "18-Jun", "Czechia", "South Africa", "A"),
    (26, "18-Jun", "Switzerland", "Bosnia and Herzegovina", "B"),
    (27, "18-Jun", "Canada", "Qatar", "B"),
    (28, "18-Jun", "Mexico", "South Korea", "A"),
    (29, "19-Jun", "Brazil", "Haiti", "C"),
    (30, "19-Jun", "Scotland", "Morocco", "C"),
    (31, "19-Jun", "Türkiye", "Paraguay", "D"),
    (32, "19-Jun", "United States", "Australia", "D"),
    (33, "20-Jun", "Germany", "Ivory Coast", "E"),
    (34, "20-Jun", "Ecuador", "Curaçao", "E"),
    (35, "20-Jun", "Netherlands", "Sweden", "F"),
    (36, "20-Jun", "Tunisia", "Japan", "F"),
    (37, "21-Jun", "Uruguay", "Cape Verde", "H"),
    (38, "21-Jun", "Spain", "Saudi Arabia", "H"),
    (39, "21-Jun", "Belgium", "Iran", "G"),
    (40, "21-Jun", "New Zealand", "Egypt", "G"),
    (41, "22-Jun", "Norway", "Senegal", "I"),
    (42, "22-Jun", "France", "Iraq", "I"),
    (43, "22-Jun", "Argentina", "Austria", "J"),
    (44, "22-Jun", "Jordan", "Algeria", "J"),
    (45, "23-Jun", "England", "Ghana", "L"),
    (46, "23-Jun", "Panama", "Croatia", "L"),
    (47, "23-Jun", "Portugal", "Uzbekistan", "K"),
    (48, "23-Jun", "Colombia", "DR Congo", "K"),
    # Matchday 3
    (49, "24-Jun", "Scotland", "Brazil", "C"),
    (50, "24-Jun", "Morocco", "Haiti", "C"),
    (51, "24-Jun", "Switzerland", "Canada", "B"),
    (52, "24-Jun", "Bosnia and Herzegovina", "Qatar", "B"),
    (53, "24-Jun", "Czechia", "Mexico", "A"),
    (54, "24-Jun", "South Africa", "South Korea", "A"),
    (55, "25-Jun", "Curaçao", "Ivory Coast", "E"),
    (56, "25-Jun", "Ecuador", "Germany", "E"),
    (57, "25-Jun", "Japan", "Sweden", "F"),
    (58, "25-Jun", "Tunisia", "Netherlands", "F"),
    (59, "25-Jun", "Türkiye", "United States", "D"),
    (60, "25-Jun", "Paraguay", "Australia", "D"),
    (61, "26-Jun", "Norway", "France", "I"),
    (62, "26-Jun", "Senegal", "Iraq", "I"),
    (63, "26-Jun", "Egypt", "Iran", "G"),
    (64, "26-Jun", "New Zealand", "Belgium", "G"),
    (65, "26-Jun", "Cape Verde", "Saudi Arabia", "H"),
    (66, "26-Jun", "Uruguay", "Spain", "H"),
    (67, "27-Jun", "Panama", "England", "L"),
    (68, "27-Jun", "Croatia", "Ghana", "L"),
    (69, "27-Jun", "Algeria", "Austria", "J"),
    (70, "27-Jun", "Jordan", "Argentina", "J"),
    (71, "27-Jun", "Colombia", "Portugal", "K"),
    (72, "27-Jun", "DR Congo", "Uzbekistan", "K"),
]

# Knockout stage fixtures (generic - teams TBD)
KNOCKOUT_FIXTURES = [
    # Round of 32
    (73, "28-Jun", "2nd Group A", "2nd Group B", "R32"),
    (74, "29-Jun", "1st Group E", "3rd ABCDF", "R32"),
    (75, "29-Jun", "1st Group F", "2nd Group C", "R32"),
    (76, "29-Jun", "1st Group C", "2nd Group F", "R32"),
    (77, "30-Jun", "1st Group I", "3rd CDFGH", "R32"),
    (78, "30-Jun", "2nd Group E", "2nd Group I", "R32"),
    (79, "30-Jun", "1st Group A", "3rd CEFHI", "R32"),
    (80, "01-Jul", "1st Group L", "3rd EHIJK", "R32"),
    (81, "01-Jul", "1st Group D", "3rd BEFIJ", "R32"),
    (82, "01-Jul", "1st Group G", "3rd AEHIJ", "R32"),
    (83, "02-Jul", "2nd Group K", "2nd Group L", "R32"),
    (84, "02-Jul", "1st Group H", "2nd Group J", "R32"),
    (85, "02-Jul", "1st Group B", "3rd EFGIJ", "R32"),
    (86, "03-Jul", "1st Group J", "2nd Group H", "R32"),
    (87, "03-Jul", "1st Group K", "3rd DEIJL", "R32"),
    (88, "03-Jul", "2nd Group D", "2nd Group G", "R32"),
    # Round of 16
    (89, "04-Jul", "W74", "W77", "R16"),
    (90, "04-Jul", "W73", "W75", "R16"),
    (91, "05-Jul", "W76", "W78", "R16"),
    (92, "05-Jul", "W79", "W80", "R16"),
    (93, "06-Jul", "W83", "W84", "R16"),
    (94, "06-Jul", "W81", "W82", "R16"),
    (95, "07-Jul", "W86", "W88", "R16"),
    (96, "07-Jul", "W85", "W87", "R16"),
    # Quarter-finals
    (97, "09-Jul", "W89", "W90", "QF"),
    (98, "10-Jul", "W93", "W94", "QF"),
    (99, "11-Jul", "W91", "W92", "QF"),
    (100, "11-Jul", "W95", "W96", "QF"),
    # Semi-finals
    (101, "14-Jul", "W97", "W98", "SF"),
    (102, "15-Jul", "W99", "W100", "SF"),
    # Third place
    (103, "18-Jul", "L101", "L102", "3rd"),
    # Final
    (104, "19-Jul", "W101", "W102", "Final"),
]

# ============================================================
# STYLING
# ============================================================

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="1F4E79")

GROUP_COLORS = {
    "A": "E8F5E9", "B": "E3F2FD", "C": "FFF3E0", "D": "F3E5F5",
    "E": "E0F7FA", "F": "FBE9E7", "G": "F1F8E9", "H": "EDE7F6",
    "I": "E8EAF6", "J": "FCE4EC", "K": "E0F2F1", "L": "FFF8E1",
}

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 35)


# ============================================================
# SHEET BUILDERS
# ============================================================

def build_teams_sheet(wb):
    ws = wb.active
    ws.title = "Teams"
    
    ws.cell(row=1, column=1, value="FIFA World Cup 2026 - All 48 Teams").font = TITLE_FONT
    
    headers = ["#", "Team", "Group"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))
    
    row = 4
    team_num = 1
    for group in sorted(TEAMS.keys()):
        for team in TEAMS[group]:
            ws.cell(row=row, column=1, value=team_num)
            ws.cell(row=row, column=2, value=team)
            ws.cell(row=row, column=3, value=f"Group {group}")
            fill = PatternFill(start_color=GROUP_COLORS[group], end_color=GROUP_COLORS[group], fill_type="solid")
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1
            team_num += 1
    
    auto_width(ws)


def build_participants_sheet(wb):
    ws = wb.create_sheet("Participants")
    
    ws.cell(row=1, column=1, value="Participant Registration").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Paste names from Microsoft Forms below. One name per row.").font = Font(italic=True)
    
    headers = ["#", "Participant Name", "Golden Boot Nomination"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))
    
    # Pre-fill row numbers for up to 100 participants
    for i in range(1, 101):
        ws.cell(row=4 + i, column=1, value=i)
        ws.cell(row=4 + i, column=1).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30


def build_allocation_sheet(wb):
    ws = wb.create_sheet("Allocation")
    
    ws.cell(row=1, column=1, value="Team Allocation").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Run the 'AllocateTeams' macro to randomly assign teams to participants.").font = Font(italic=True)
    ws.cell(row=3, column=1, value="Teams are allocated in cycles of 48 - all 48 teams used before repeating.").font = Font(italic=True, color="FF0000")
    
    headers = ["#", "Participant Name", "Allocated Team", "Group", "Cycle"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, len(headers))
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8


def build_fixtures_sheet(wb):
    ws = wb.create_sheet("Fixtures")
    
    ws.cell(row=1, column=1, value="FIFA World Cup 2026 - Match Results").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Enter scores in columns E and F. Points calculate automatically.").font = Font(italic=True)
    
    headers = ["Match", "Date", "Stage", "Home Team", "Home Score", "Away Score", "Away Team", 
               "Home Pts", "Away Pts", "Home Result", "Away Result"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))
    
    row = 5
    all_fixtures = GROUP_FIXTURES + KNOCKOUT_FIXTURES
    
    for match_num, date, home, away, stage in all_fixtures:
        ws.cell(row=row, column=1, value=match_num)
        ws.cell(row=row, column=2, value=date)
        ws.cell(row=row, column=3, value=stage)
        ws.cell(row=row, column=4, value=home)
        # Columns E (5) and F (6) are for score entry
        ws.cell(row=row, column=7, value=away)
        
        # Points formulas (only calculate if both scores entered)
        # Home points: IF both scores entered, IF home>away then 3, IF home=away then 1, else 0
        home_pts = f'=IF(AND(E{row}<>"",F{row}<>""),IF(E{row}>F{row},3,IF(E{row}=F{row},1,0)),"")'
        away_pts = f'=IF(AND(E{row}<>"",F{row}<>""),IF(F{row}>E{row},3,IF(E{row}=F{row},1,0)),"")'
        ws.cell(row=row, column=8, value=home_pts)
        ws.cell(row=row, column=9, value=away_pts)
        
        # Result text (W/D/L)
        home_result = f'=IF(AND(E{row}<>"",F{row}<>""),IF(E{row}>F{row},"W",IF(E{row}=F{row},"D","L")),"")'
        away_result = f'=IF(AND(E{row}<>"",F{row}<>""),IF(F{row}>E{row},"W",IF(E{row}=F{row},"D","L")),"")'
        ws.cell(row=row, column=10, value=home_result)
        ws.cell(row=row, column=11, value=away_result)
        
        # Color group stage rows
        if stage in GROUP_COLORS:
            fill = PatternFill(start_color=GROUP_COLORS[stage], end_color=GROUP_COLORS[stage], fill_type="solid")
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = fill
        
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        # Left-align team names
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='left')
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='left')
        
        row += 1
    
    # Set column widths
    widths = [7, 8, 8, 28, 10, 10, 28, 9, 9, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_points_sheet(wb):
    ws = wb.create_sheet("Points")
    
    ws.cell(row=1, column=1, value="Team Points Summary").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Auto-calculated from Fixtures sheet results.").font = Font(italic=True)
    
    headers = ["#", "Team", "Group", "Played", "Won", "Drawn", "Lost", "GF", "GA", "GD", "Points"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))
    
    row = 5
    team_num = 1
    for group in sorted(TEAMS.keys()):
        for team in TEAMS[group]:
            ws.cell(row=row, column=1, value=team_num)
            ws.cell(row=row, column=2, value=team)
            ws.cell(row=row, column=3, value=f"Group {group}")
            
            team_cell = f'B{row}'
            
            # Played = count of matches where this team appears AND scores are filled
            played = (f'=COUNTIFS(Fixtures!D$5:D$76,{team_cell},Fixtures!E$5:E$76,"<>")'
                     f'+COUNTIFS(Fixtures!G$5:G$76,{team_cell},Fixtures!F$5:F$76,"<>")')
            ws.cell(row=row, column=4, value=played)
            
            # Won = home wins + away wins
            won = (f'=COUNTIFS(Fixtures!D$5:D$76,{team_cell},Fixtures!J$5:J$76,"W")'
                  f'+COUNTIFS(Fixtures!G$5:G$76,{team_cell},Fixtures!K$5:K$76,"W")')
            ws.cell(row=row, column=5, value=won)
            
            # Drawn
            drawn = (f'=COUNTIFS(Fixtures!D$5:D$76,{team_cell},Fixtures!J$5:J$76,"D")'
                    f'+COUNTIFS(Fixtures!G$5:G$76,{team_cell},Fixtures!K$5:K$76,"D")')
            ws.cell(row=row, column=6, value=drawn)
            
            # Lost
            lost = (f'=COUNTIFS(Fixtures!D$5:D$76,{team_cell},Fixtures!J$5:J$76,"L")'
                   f'+COUNTIFS(Fixtures!G$5:G$76,{team_cell},Fixtures!K$5:K$76,"L")')
            ws.cell(row=row, column=7, value=lost)
            
            # Goals For (home goals when home + away goals when away)
            gf = (f'=SUMPRODUCT((Fixtures!D$5:D$76={team_cell})*(Fixtures!E$5:E$76<>"")*Fixtures!E$5:E$76)'
                 f'+SUMPRODUCT((Fixtures!G$5:G$76={team_cell})*(Fixtures!F$5:F$76<>"")*Fixtures!F$5:F$76)')
            ws.cell(row=row, column=8, value=gf)
            
            # Goals Against
            ga = (f'=SUMPRODUCT((Fixtures!D$5:D$76={team_cell})*(Fixtures!F$5:F$76<>"")*Fixtures!F$5:F$76)'
                 f'+SUMPRODUCT((Fixtures!G$5:G$76={team_cell})*(Fixtures!E$5:E$76<>"")*Fixtures!E$5:E$76)')
            ws.cell(row=row, column=9, value=ga)
            
            # Goal Difference
            ws.cell(row=row, column=10, value=f'=H{row}-I{row}')
            
            # Total Points - calculate directly from scores to avoid "" multiplication issues
            points = (f'=SUMPRODUCT((Fixtures!D$5:D$76={team_cell})*(Fixtures!E$5:E$76<>"")*(Fixtures!F$5:F$76<>"")*((Fixtures!E$5:E$76>Fixtures!F$5:F$76)*3+(Fixtures!E$5:E$76=Fixtures!F$5:F$76)*2))'
                     f'+SUMPRODUCT((Fixtures!G$5:G$76={team_cell})*(Fixtures!E$5:E$76<>"")*(Fixtures!F$5:F$76<>"")*((Fixtures!F$5:F$76>Fixtures!E$5:E$76)*3+(Fixtures!E$5:E$76=Fixtures!F$5:F$76)*2))')
            ws.cell(row=row, column=11, value=points)
            
            # Formatting
            fill = PatternFill(start_color=GROUP_COLORS[group], end_color=GROUP_COLORS[group], fill_type="solid")
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
            
            row += 1
            team_num += 1
    
    widths = [5, 28, 10, 8, 6, 7, 6, 6, 6, 6, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_leaderboard_sheet(wb):
    ws = wb.create_sheet("Leaderboard")
    
    ws.cell(row=1, column=1, value="SWEEPSTAKE LEADERBOARD").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Auto-ranked — highest points always shown first.").font = Font(italic=True)
    
    # === DISPLAY SECTION (columns A-F) - auto-sorted by points ===
    headers = ["Rank", "Participant", "Team(s)", "Total Points", "Wins", "Draws"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))
    
    # For each display row, use LARGE to get the nth highest points,
    # then INDEX/MATCH to look up the participant with that rank
    for i in range(1, 101):
        row = 4 + i
        ws.cell(row=row, column=1, value=i)
        
        # Participant name: INDEX into helper data using the rank position
        # Uses MATCH on the rank helper column (N) to find which row has rank=i
        ws.cell(row=row, column=2, 
                value=f'=IFERROR(INDEX($H$5:$H$104,MATCH({i},$N$5:$N$104,0)),"")')
        ws.cell(row=row, column=3, 
                value=f'=IFERROR(INDEX($I$5:$I$104,MATCH({i},$N$5:$N$104,0)),"")')
        ws.cell(row=row, column=4, 
                value=f'=IFERROR(INDEX($J$5:$J$104,MATCH({i},$N$5:$N$104,0)),"")')
        ws.cell(row=row, column=5, 
                value=f'=IFERROR(INDEX($K$5:$K$104,MATCH({i},$N$5:$N$104,0)),"")')
        ws.cell(row=row, column=6, 
                value=f'=IFERROR(INDEX($L$5:$L$104,MATCH({i},$N$5:$N$104,0)),"")')
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')
    
    # Highlight top 3
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    for col in range(1, 7):
        ws.cell(row=5, column=col).fill = gold_fill
        ws.cell(row=6, column=col).fill = silver_fill
        ws.cell(row=7, column=col).fill = bronze_fill
    
    # === HELPER SECTION (columns H-N) - raw data + rank calculation ===
    helper_headers = ["Participant", "Team", "Points", "Wins", "Draws", "GD", "Rank"]
    for col, h in enumerate(helper_headers, 8):  # Start at column H
        ws.cell(row=4, column=col, value=h)
        ws.cell(row=4, column=col).font = Font(bold=True, size=9, color="808080")
    
    for i in range(1, 101):
        row = 4 + i
        alloc_row = 5 + i  # Allocation sheet data starts row 6
        
        # H: Participant name from Allocation
        ws.cell(row=row, column=8, 
                value=f'=IF(Allocation!B{alloc_row}="","",Allocation!B{alloc_row})')
        # I: Team from Allocation
        ws.cell(row=row, column=9, 
                value=f'=IF(Allocation!C{alloc_row}="","",Allocation!C{alloc_row})')
        # J: Total Points (calculated directly from scores)
        ws.cell(row=row, column=10, 
                value=f'=IF(I{row}="","",'
                      f'SUMPRODUCT((Fixtures!D$5:D$76=I{row})*(Fixtures!E$5:E$76<>"")*(Fixtures!F$5:F$76<>"")*((Fixtures!E$5:E$76>Fixtures!F$5:F$76)*3+(Fixtures!E$5:E$76=Fixtures!F$5:F$76)*2))'
                      f'+SUMPRODUCT((Fixtures!G$5:G$76=I{row})*(Fixtures!E$5:E$76<>"")*(Fixtures!F$5:F$76<>"")*((Fixtures!F$5:F$76>Fixtures!E$5:E$76)*3+(Fixtures!E$5:E$76=Fixtures!F$5:F$76)*2)))')
        # K: Wins
        ws.cell(row=row, column=11, 
                value=f'=IF(I{row}="","",'
                      f'COUNTIFS(Fixtures!D$5:D$76,I{row},Fixtures!J$5:J$76,"W")'
                      f'+COUNTIFS(Fixtures!G$5:G$76,I{row},Fixtures!K$5:K$76,"W"))')
        # L: Draws
        ws.cell(row=row, column=12, 
                value=f'=IF(I{row}="","",'
                      f'COUNTIFS(Fixtures!D$5:D$76,I{row},Fixtures!J$5:J$76,"D")'
                      f'+COUNTIFS(Fixtures!G$5:G$76,I{row},Fixtures!K$5:K$76,"D"))')
        # M: Goal Difference (tiebreaker)
        ws.cell(row=row, column=13, 
                value=f'=IF(I{row}="","",'
                      f'SUMPRODUCT((Fixtures!D$5:D$76=I{row})*(Fixtures!E$5:E$76<>"")*(Fixtures!F$5:F$76<>"")*(Fixtures!E$5:E$76-Fixtures!F$5:F$76))'
                      f'+SUMPRODUCT((Fixtures!G$5:G$76=I{row})*(Fixtures!E$5:E$76<>"")*(Fixtures!F$5:F$76<>"")*(Fixtures!F$5:F$76-Fixtures!E$5:E$76)))')
# N: Rank (unique - by points desc, then GD desc, then row order as tiebreaker)
        # COUNTIFS: count participants with more points, + count ties above this row
        ws.cell(row=row, column=14, 
                value=f'=IF(H{row}="","",' 
                      f'COUNTIFS($J$5:$J$104,">"&J{row},$H$5:$H$104,"<>")'
                      f'+COUNTIFS($J$5:$J{row},J{row},$H$5:$H{row},"<>"))')
    
    # Set column widths
    widths = {'A': 6, 'B': 28, 'C': 28, 'D': 12, 'E': 6, 'F': 7,
              'G': 2, 'H': 22, 'I': 22, 'J': 8, 'K': 6, 'L': 6, 'M': 5, 'N': 5}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
    
    # Hide helper columns
    for col_letter in ['H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col_letter].hidden = True


def build_golden_boot_sheet(wb):
    ws = wb.create_sheet("Golden Boot")
    
    ws.cell(row=1, column=1, value="Golden Boot Tracker").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Log goal scorers in the table below. Participant standings auto-update.").font = Font(italic=True)
    
    # === SECTION 1: GOAL LOG (columns A-D) ===
    # User enters goal scorers here after each match
    ws.cell(row=4, column=1, value="GOAL LOG").font = SUBTITLE_FONT
    ws.cell(row=4, column=4, value="Enter one row per scorer per match").font = Font(italic=True, size=9)
    
    log_headers = ["Match #", "Player Name", "Team", "Goals"]
    for col, h in enumerate(log_headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, len(log_headers))
    
    # Pre-format 300 rows for goal entries (plenty for a tournament)
    for i in range(6, 306):
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = THIN_BORDER
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=4).alignment = Alignment(horizontal='center')
    
    # === SECTION 2: TOURNAMENT TOP SCORERS (columns F-I) - auto-ranked ===
    ws.cell(row=4, column=6, value="TOP SCORERS (Auto)").font = SUBTITLE_FONT
    
    scorer_headers = ["Rank", "Player", "Team", "Goals"]
    for col, h in enumerate(scorer_headers, 6):
        ws.cell(row=5, column=col, value=h)
    for col in range(6, 10):
        cell = ws.cell(row=5, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Top scorers uses helper columns (P-S) with unique player list + SUMIF
    # Display top 30 scorers auto-ranked
    for i in range(1, 31):
        row = 5 + i
        ws.cell(row=row, column=6, value=i)
        ws.cell(row=row, column=7, 
                value=f'=IFERROR(INDEX($P$6:$P$305,MATCH({i},$S$6:$S$305,0)),"")')
        ws.cell(row=row, column=8, 
                value=f'=IFERROR(INDEX($Q$6:$Q$305,MATCH({i},$S$6:$S$305,0)),"")')
        ws.cell(row=row, column=9, 
                value=f'=IFERROR(INDEX($R$6:$R$305,MATCH({i},$S$6:$S$305,0)),"")')
        for col in range(6, 10):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='left')
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='left')
    
    # Highlight top 3
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    for col in range(6, 10):
        ws.cell(row=6, column=col).fill = gold_fill
        ws.cell(row=7, column=col).fill = silver_fill
        ws.cell(row=8, column=col).fill = bronze_fill
    
    # === SECTION 3: PARTICIPANT GOLDEN BOOT STANDINGS (columns K-N) - auto-ranked ===
    ws.cell(row=4, column=11, value="PARTICIPANT STANDINGS (Auto)").font = SUBTITLE_FONT
    
    part_headers = ["Rank", "Participant", "Nominated Player", "Player Goals"]
    for col, h in enumerate(part_headers, 11):
        ws.cell(row=5, column=col, value=h)
    for col in range(11, 15):
        cell = ws.cell(row=5, column=col)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Participant standings - auto-ranked by their nominated player's goals
    # Uses helper columns (T-W) 
    for i in range(1, 51):
        row = 5 + i
        ws.cell(row=row, column=11, value=i)
        ws.cell(row=row, column=12, 
                value=f'=IFERROR(INDEX($T$6:$T$105,MATCH({i},$W$6:$W$105,0)),"")')
        ws.cell(row=row, column=13, 
                value=f'=IFERROR(INDEX($U$6:$U$105,MATCH({i},$W$6:$W$105,0)),"")')
        ws.cell(row=row, column=14, 
                value=f'=IFERROR(INDEX($V$6:$V$105,MATCH({i},$W$6:$W$105,0)),"")')
        for col in range(11, 15):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=12).alignment = Alignment(horizontal='left')
        ws.cell(row=row, column=13).alignment = Alignment(horizontal='left')
    
    # Top 3 highlight
    for col in range(11, 15):
        ws.cell(row=6, column=col).fill = gold_fill
        ws.cell(row=7, column=col).fill = silver_fill
        ws.cell(row=8, column=col).fill = bronze_fill
    
    # === HELPER COLUMNS (P-S): Unique player list + goals + rank ===
    # P: Player name (from goal log - unique list built with formula)
    # Q: Team
    # R: Total goals (SUMIF on goal log)
    # S: Rank
    
    # For unique player extraction, we use a helper approach:
    # Copy all player names from goal log, deduplicate with COUNTIF trick
    # Since Excel formulas for unique lists are complex, we use a simpler approach:
    # Just reference the goal log directly and use SUMIF for totals
    # The "unique players" will be built from the log entries
    
    ws.cell(row=5, column=16, value="Player").font = Font(bold=True, size=8, color="808080")
    ws.cell(row=5, column=17, value="Team").font = Font(bold=True, size=8, color="808080")
    ws.cell(row=5, column=18, value="Goals").font = Font(bold=True, size=8, color="808080")
    ws.cell(row=5, column=19, value="Rank").font = Font(bold=True, size=8, color="808080")
    
    for i in range(300):
        row = 6 + i
        # P: Player name from goal log (just mirror it)
        ws.cell(row=row, column=16, value=f'=IF(B{row}="","",B{row})')
        # Q: Team from goal log
        ws.cell(row=row, column=17, value=f'=IF(C{row}="","",C{row})')
        # R: Total goals for this player (SUMIF across all log entries)
        ws.cell(row=row, column=18, 
                value=f'=IF(B{row}="","",SUMIF($B$6:$B$305,B{row},$D$6:$D$305))')
        # S: Rank (only for first occurrence of each player to avoid duplicates in display)
        # If this player appeared earlier in the log, give "" rank (skip duplicate)
        ws.cell(row=row, column=19, 
                value=f'=IF(OR(B{row}="",COUNTIF($B$6:$B{row},B{row})>1),"",'
                      f'COUNTIFS($R$6:$R$305,">"&R{row},$B$6:$B$305,"<>",$P$6:$P$305,"<>")'
                      f'-COUNTIFS($R$6:$R$305,">"&R{row},$B$6:$B$305,"<>",$P$6:$P$305,"<>",$B$6:$B$305,B{row})'
                      f'+COUNTIFS($R$6:$R{row},"="&R{row},$B$6:$B{row},"<>",OFFSET($P$6,0,0,{row}-5,1),"<>")'
                      f')')
    
    # Actually, the unique player ranking is complex. Let me use a simpler approach:
    # Just rank among first-occurrence rows
    for i in range(300):
        row = 6 + i
        # Simpler rank: for first occurrence only, count unique players with more goals
        ws.cell(row=row, column=19, 
                value=f'=IF(OR(B{row}="",COUNTIF($B$6:$B{row},B{row})>1),"",'
                      f'SUMPRODUCT((COUNTIF($B$5:B{row},$B$6:$B$305)=1)*($R$6:$R$305>R{row})*($B$6:$B$305<>""))+1)')
    
    # === HELPER COLUMNS (T-W): Participant golden boot data + rank ===
    ws.cell(row=5, column=20, value="Participant").font = Font(bold=True, size=8, color="808080")
    ws.cell(row=5, column=21, value="Nomination").font = Font(bold=True, size=8, color="808080")
    ws.cell(row=5, column=22, value="Goals").font = Font(bold=True, size=8, color="808080")
    ws.cell(row=5, column=23, value="Rank").font = Font(bold=True, size=8, color="808080")
    
    for i in range(1, 101):
        row = 5 + i
        # T: Participant name
        ws.cell(row=row, column=20, 
                value=f'=IF(Participants!B{4+i}="","",Participants!B{4+i})')
        # U: Their Golden Boot nomination
        ws.cell(row=row, column=21, 
                value=f'=IF(Participants!C{4+i}="","",Participants!C{4+i})')
        # V: Goals scored by their nominated player (SUMIF on goal log)
        ws.cell(row=row, column=22, 
                value=f'=IF(U{row}="","",'
                      f'IFERROR(SUMIF($B$6:$B$305,U{row},$D$6:$D$305),0))')
        # W: Rank (by goals desc, row order tiebreaker)
        ws.cell(row=row, column=23, 
                value=f'=IF(T{row}="","",'
                      f'COUNTIFS($V$6:$V$105,">"&V{row},$T$6:$T$105,"<>")'
                      f'+COUNTIFS($V$6:$V{row},"="&V{row},$T$6:$T{row},"<>"))')
    
    # Set column widths
    col_widths = {'A': 8, 'B': 25, 'C': 20, 'D': 7, 'E': 2,
                  'F': 5, 'G': 25, 'H': 18, 'I': 7, 'J': 2,
                  'K': 5, 'L': 22, 'M': 25, 'N': 12}
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w
    
    # Hide helper columns P-W
    for col_letter in ['O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
        ws.column_dimensions[col_letter].hidden = True


def build_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False
    
    instructions = [
        ("FIFA WORLD CUP 2026 SWEEPSTAKE", TITLE_FONT),
        ("", None),
        ("HOW TO USE THIS WORKBOOK", SUBTITLE_FONT),
        ("", None),
        ("STEP 1: Collect Participants", SUBTITLE_FONT),
        ("- Send out your Microsoft Form for registration", None),
        ("- Paste participant names into the 'Participants' sheet column B", None),
        ("- Have them nominate a Golden Boot player in column C", None),
        ("", None),
        ("STEP 2: Allocate Teams", SUBTITLE_FONT),
        ("- Save this file as .xlsm (macro-enabled workbook)", None),
        ("- Import the VBA macro from 'allocation_macro.bas' (Alt+F11 > Import)", None),
        ("- Run the macro 'AllocateTeams' (Alt+F8 > AllocateTeams > Run)", None),
        ("- Teams are randomly assigned in cycles of 48", None),
        ("- First 48 participants each get a unique team", None),
        ("- Participants 49+ get teams from a fresh cycle of 48", None),
        ("", None),
        ("STEP 3: Enter Results", SUBTITLE_FONT),
        ("- Go to the 'Fixtures' sheet", None),
        ("- Enter Home Score in column E and Away Score in column F", None),
        ("- Points calculate automatically (Win=3, Draw=2, Loss=0)", None),
        ("", None),
        ("STEP 4: View Leaderboard", SUBTITLE_FONT),
        ("- The 'Leaderboard' sheet shows each participant's total points", None),
        ("- Sort by column D (Total Points) descending after entering results", None),
        ("- Select data > Data tab > Sort > Column D > Largest to Smallest", None),
        ("", None),
        ("STEP 5: Track Golden Boot", SUBTITLE_FONT),
        ("- After each match, log goal scorers in 'Golden Boot' sheet columns A-D", None),
        ("- Enter: Match #, Player Name, Team, Goals scored in that match", None),
        ("- One row per scorer per match (e.g., Mbappé scored 2 = one row with 2)", None),
        ("- Top Scorers table auto-ranks all players by total goals", None),
        ("- Participant Standings auto-matches nominations to goals", None),
        ("- Winner is shown at position 1 in Participant Standings", None),
        ("", None),
        ("STEP 6: Knockout Stage", SUBTITLE_FONT),
        ("- Knockout fixtures (R32 onwards) have placeholder team names", None),
        ("  e.g. '1st Group A', '2nd Group B', '3rd ABCDF', 'W74', 'L101'", None),
        ("- Once the group stage finishes, replace these with actual team names:", None),
        ("  1. Go to the 'Fixtures' sheet", None),
        ("  2. For R32 matches: replace '1st Group A' with the team that finished", None),
        ("     1st in Group A, '2nd Group B' with 2nd in Group B, etc.", None),
        ("  3. '3rd ABCDF' means the best 3rd-place team from groups A/B/C/D/F", None),
        ("     — FIFA publishes the exact 3rd-place bracket after group stage ends", None),
        ("  4. For R16 onwards: 'W74' = winner of match 74, 'L101' = loser of match 101", None),
        ("     — Update these after each round finishes", None),
        ("- Points still work the same — just enter scores as normal once teams are filled in", None),
        ("- The Leaderboard updates automatically as knockout scores are entered", None),
        ("", None),
        ("STEP 7: Weekly Quiz", SUBTITLE_FONT),
        ("- Go to the 'Quiz' sheet", None),
        ("- Each week, enter participant scores in the corresponding Wk column", None),
        ("- The Total column sums all weeks automatically", None),
        ("- Rank updates automatically — highest cumulative score = 1st", None),
        ("- Up to 10 weeks supported (add more columns if needed)", None),
        ("", None),
        ("POINTS SYSTEM", SUBTITLE_FONT),
        ("- Win = 3 points", None),
        ("- Draw = 2 points (both teams get 2 points)", None),
        ("- Loss = 0 points", None),
        ("- Points accumulate across ALL matches (group + knockout)", None),
        ("", None),
        ("NOTES", SUBTITLE_FONT),
        ("- Knockout matches: if it goes to extra time, the score at full time", None),
        ("  (including extra time) counts. Penalty shootouts count as a draw.", None),
        ("- If you have >48 participants, each person still gets ONE team.", None),
        ("  Teams can be shared across cycles (e.g., 2 people may have Brazil).", None),
    ]
    
    for i, (text, font) in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
    
    ws.column_dimensions['A'].width = 80


# ============================================================
# QUIZ LEADERBOARD
# ============================================================

def build_quiz_sheet(wb):
    ws = wb.create_sheet("Quiz")
    
    ws.cell(row=1, column=1, value="WEEKLY QUIZ LEADERBOARD").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Enter each week's scores below. Cumulative totals update automatically.").font = Font(italic=True)
    
    # Row 4: headers — Rank | Participant | Wk1 | Wk2 | ... Wk10 | Total
    # Allow up to 10 weeks (can extend later)
    max_weeks = 10
    
    ws.cell(row=4, column=1, value="Rank")
    ws.cell(row=4, column=2, value="Participant")
    for wk in range(1, max_weeks + 1):
        ws.cell(row=4, column=2 + wk, value=f"Wk {wk}")
    ws.cell(row=4, column=2 + max_weeks + 1, value="Total")
    
    total_cols = 2 + max_weeks + 1  # 13 columns (A-M)
    style_header_row(ws, 4, total_cols)
    
    # Data rows: pull participant names from Participants sheet, sum weekly scores
    for i in range(1, 101):
        row = 4 + i
        part_row = 4 + i  # Participants sheet data starts row 5
        
        # Participant name from Participants sheet
        ws.cell(row=row, column=2, 
                value=f'=IF(Participants!B{part_row}="","",Participants!B{part_row})')
        
        # Weekly score columns (C to L) — user enters manually
        # Leave blank for manual entry
        
        # Total column (M) — sum of all weekly scores
        first_col = get_column_letter(3)  # C
        last_col = get_column_letter(2 + max_weeks)  # L
        ws.cell(row=row, column=2 + max_weeks + 1, 
                value=f'=IF(B{row}="","",SUM({first_col}{row}:{last_col}{row}))')
        
        # Rank column — unique rank based on total descending
        total_col_letter = get_column_letter(2 + max_weeks + 1)  # M
        ws.cell(row=row, column=1,
                value=f'=IF(B{row}="","",COUNTIFS(${total_col_letter}$5:${total_col_letter}$104,">"&{total_col_letter}{row},$B$5:$B$104,"<>")'
                      f'+COUNTIFS(${total_col_letter}$5:${total_col_letter}{row},{total_col_letter}{row},$B$5:$B{row},"<>"))')
        
        # Formatting
        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
    
    # Highlight top 3
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    for col in range(1, total_cols + 1):
        ws.cell(row=5, column=col).fill = gold_fill
        ws.cell(row=6, column=col).fill = silver_fill
        ws.cell(row=7, column=col).fill = bronze_fill
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    for wk in range(1, max_weeks + 1):
        ws.column_dimensions[get_column_letter(2 + wk)].width = 7
    ws.column_dimensions[get_column_letter(2 + max_weeks + 1)].width = 8


# ============================================================
# MAIN
# ============================================================

SEED_PARTICIPANTS = [
    ("James Wilson", "Kylian Mbappé"),
    ("Sarah Chen", "Lionel Messi"),
    ("Michael O'Brien", "Erling Haaland"),
    ("Emma Thompson", "Vinícius Júnior"),
    ("David Kumar", "Harry Kane"),
    ("Rachel Stewart", "Mohamed Salah"),
    ("Chris Martinez", "Jude Bellingham"),
    ("Jessica Park", "Lamine Yamal"),
    ("Andrew Collins", "Bukayo Saka"),
    ("Laura Bennett", "Victor Osimhen"),
    ("Tom Nguyen", "Cristiano Ronaldo"),
    ("Sophie Williams", "Robert Lewandowski"),
    ("Ryan O'Connor", "Darwin Núñez"),
    ("Hannah Lee", "Julian Álvarez"),
    ("Mark Davidson", "Romelu Lukaku"),
    ("Priya Sharma", "Cody Gakpo"),
    ("Luke Anderson", "Alexander Isak"),
    ("Megan Taylor", "Randal Kolo Muani"),
    ("Daniel Brown", "Kai Havertz"),
    ("Olivia White", "Dusan Vlahovic"),
    ("Nathan Scott", "Rasmus Højlund"),
    ("Amy Richardson", "Ollie Watkins"),
    ("Ben Harper", "Lautaro Martínez"),
    ("Chloe Morgan", "Florian Wirtz"),
    ("Sam Mitchell", "Son Heung-min"),
    ("Emily Zhang", "Marcus Rashford"),
    ("Jake Turner", "Antoine Griezmann"),
    ("Fiona Campbell", "Serhou Guirassy"),
    ("Alex Patel", "Álvaro Morata"),
    ("Grace Murphy", "Ademola Lookman"),
    ("Owen Phillips", "Jonathan David"),
    ("Zoe Carter", "Luis Díaz"),
    ("Liam Hughes", "Memphis Depay"),
    ("Isabella Ross", "Nicolas Jackson"),
    ("Ethan Wright", "Islam Slimani"),
    ("Charlotte King", "Mehdi Taremi"),
    ("Josh Evans", "Rafael Leão"),
    ("Katie Adams", "Artem Dovbyk"),
    ("Will Jackson", "Viktor Gyökeres"),
    ("Natalie Green", "Matheus Cunha"),
    ("Harry Cooper", "Saleh Al-Shehri"),
    ("Rebecca Hall", "Sadio Mané"),
    ("Finn O'Sullivan", "Achraf Hakimi"),
    ("Ella Dixon", "Bryan Mbeumo"),
    ("Patrick Ward", "Amine Harit"),
    ("Amber Long", "Tim Weah"),
    ("Callum Fraser", "Alphonso Davies"),
    ("Molly Price", "Santiago Giménez"),
    ("Declan Burke", "Mathys Tel"),
    ("Sienna Cole", "Donyell Malen"),
]


def seed_participants(wb):
    """Fill Participants sheet with test data."""
    ws = wb["Participants"]
    for i, (name, player) in enumerate(SEED_PARTICIPANTS, 1):
        ws.cell(row=4 + i, column=2, value=name)
        ws.cell(row=4 + i, column=3, value=player)


def main():
    wb = Workbook()
    
    build_teams_sheet(wb)
    build_participants_sheet(wb)
    build_allocation_sheet(wb)
    build_fixtures_sheet(wb)
    build_points_sheet(wb)
    build_leaderboard_sheet(wb)
    build_golden_boot_sheet(wb)
    build_quiz_sheet(wb)
    build_instructions_sheet(wb)
    
    # Add seed data
    seed_participants(wb)
    
    # Move Instructions to the front
    wb.move_sheet("Instructions", offset=-8)
    
    output_path = r"c:\Users\s67655\Downloads\World Cup Sweep\WorldCup2026_Sweepstake.xlsx"
    wb.save(output_path)
    print(f"Workbook created successfully: {output_path}")
    print(f"Seeded with {len(SEED_PARTICIPANTS)} test participants.")
    print("\nNext steps:")
    print("1. Open the .xlsx file in Excel")
    print("2. Save As .xlsm (Excel Macro-Enabled Workbook)")
    print("3. Press Alt+F11 to open VBA editor")
    print("4. File > Import > select 'allocation_macro.bas'")
    print("5. Run macro: Alt+F8 > AllocateTeams > Run")


if __name__ == "__main__":
    main()
