"""
Export World Cup 2026 Sweepstake data from Excel to JSON for GitHub Pages.

Usage:
    python export_to_web.py

Reads: WorldCup2026_Sweepstake.xlsm
Writes: docs/data.json

Run this after updating scores, then commit & push to update GitHub Pages.
"""

import json
import os
from openpyxl import load_workbook
from datetime import datetime

WORKBOOK_PATH = "WorldCup2026_Sweepstake.xlsm"
OUTPUT_DIR = "docs"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "data.json")


def export():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)

    data = {
        "lastUpdated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "teams": extract_teams(wb),
        "fixtures": extract_fixtures(wb),
        "groups": build_group_standings(wb),
        "leaderboard": extract_leaderboard(wb),
    }

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Exported to {OUTPUT_PATH}")
    print(f"  Teams: {len(data['teams'])}")
    print(f"  Fixtures: {len(data['fixtures'])}")
    print(f"  Leaderboard entries: {len(data['leaderboard'])}")
    print(f"\nCommit & push to update GitHub Pages!")


def extract_teams(wb):
    ws = wb["Teams"]
    teams = []
    for row in range(4, 52):  # rows 4-51 = 48 teams
        team = ws.cell(row=row, column=2).value
        group = ws.cell(row=row, column=3).value
        if team:
            teams.append({"name": team, "group": group or ""})
    return teams


def extract_fixtures(wb):
    ws = wb["Fixtures"]
    fixtures = []
    for row in range(5, 109):  # rows 5-108 = 104 matches
        match_num = ws.cell(row=row, column=1).value
        if match_num is None:
            break
        date = ws.cell(row=row, column=2).value
        stage = ws.cell(row=row, column=3).value
        home = ws.cell(row=row, column=4).value
        home_score = ws.cell(row=row, column=5).value
        away_score = ws.cell(row=row, column=6).value
        away = ws.cell(row=row, column=7).value

        fixtures.append({
            "match": int(match_num) if match_num else None,
            "date": str(date) if date else "",
            "stage": stage or "",
            "home": home or "",
            "away": away or "",
            "homeScore": int(home_score) if home_score is not None and home_score != "" else None,
            "awayScore": int(away_score) if away_score is not None and away_score != "" else None,
        })
    return fixtures


def build_group_standings(wb):
    """Build group tables from fixture results, showing all teams even with no results."""
    ws = wb["Fixtures"]
    ws_teams = wb["Teams"]

    # Initialise all teams from the Teams sheet
    team_stats = {}
    for row in range(4, 52):  # rows 4-51 = 48 teams
        team = ws_teams.cell(row=row, column=2).value
        group_name = ws_teams.cell(row=row, column=3).value  # e.g. "Group A"
        if team and group_name:
            group_letter = group_name.replace("Group ", "")
            team_stats[team] = {"team": team, "group": group_letter, "p": 0, "w": 0, "d": 0, "l": 0, "gf": 0, "ga": 0, "pts": 0}

    # Update stats from fixture results
    for row in range(5, 77):  # group stage = matches 1-72
        stage = ws.cell(row=row, column=3).value
        home = ws.cell(row=row, column=4).value
        away = ws.cell(row=row, column=7).value
        home_score = ws.cell(row=row, column=5).value
        away_score = ws.cell(row=row, column=6).value

        if home_score is None or away_score is None or home_score == "" or away_score == "":
            continue

        home_score = int(home_score)
        away_score = int(away_score)

        # Home
        team_stats[home]["p"] += 1
        team_stats[home]["gf"] += home_score
        team_stats[home]["ga"] += away_score
        # Away
        team_stats[away]["p"] += 1
        team_stats[away]["gf"] += away_score
        team_stats[away]["ga"] += home_score

        if home_score > away_score:
            team_stats[home]["w"] += 1
            team_stats[home]["pts"] += 3
            team_stats[away]["l"] += 1
        elif home_score == away_score:
            team_stats[home]["d"] += 1
            team_stats[home]["pts"] += 1
            team_stats[away]["d"] += 1
            team_stats[away]["pts"] += 1
        else:
            team_stats[away]["w"] += 1
            team_stats[away]["pts"] += 3
            team_stats[home]["l"] += 1

    # Group by group letter
    groups = {}
    for team, stats in team_stats.items():
        g = stats["group"]
        if g not in groups:
            groups[g] = []
        groups[g].append(stats)

    # Sort each group by pts desc, then gd desc, then gf desc
    for g in groups:
        groups[g].sort(key=lambda x: (-x["pts"], -(x["gf"] - x["ga"]), -x["gf"]))

    return groups


def extract_leaderboard(wb):
    ws = wb["Allocation"]
    ws_fix = wb["Fixtures"]

    # Build team -> points from fixtures
    team_points = {}
    team_wins = {}
    team_draws = {}

    for row in range(5, 109):
        home = ws_fix.cell(row=row, column=4).value
        away = ws_fix.cell(row=row, column=7).value
        home_score = ws_fix.cell(row=row, column=5).value
        away_score = ws_fix.cell(row=row, column=6).value

        if home_score is None or away_score is None or home_score == "" or away_score == "":
            continue

        home_score = int(home_score)
        away_score = int(away_score)

        for t in [home, away]:
            if t not in team_points:
                team_points[t] = 0
                team_wins[t] = 0
                team_draws[t] = 0

        if home_score > away_score:
            team_points[home] += 3
            team_wins[home] += 1
        elif home_score == away_score:
            team_points[home] += 1
            team_points[away] += 1
            team_draws[home] += 1
            team_draws[away] += 1
        else:
            team_points[away] += 3
            team_wins[away] += 1

    # Read allocation
    entries = []
    for row in range(6, 200):
        participant = ws.cell(row=row, column=2).value
        team = ws.cell(row=row, column=3).value
        if not participant or not team:
            break
        pts = team_points.get(team, 0)
        wins = team_wins.get(team, 0)
        draws = team_draws.get(team, 0)
        entries.append({
            "participant": str(participant),
            "team": str(team),
            "points": pts,
            "wins": wins,
            "draws": draws,
        })

    # Sort by points desc, then wins desc
    entries.sort(key=lambda x: (-x["points"], -x["wins"]))

    # Add rank
    for i, e in enumerate(entries):
        e["rank"] = i + 1

    return entries


if __name__ == "__main__":
    export()
