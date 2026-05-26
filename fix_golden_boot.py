"""Fix Golden Boot ranking formula to handle same player in multiple rows."""
from openpyxl import load_workbook


def fix_helper_columns(ws):
    """Redesign columns P-S to use simple COUNTIFS (no SUMPRODUCT/array formulas).
    
    P = player name (first occurrence only, "" for duplicates)
    Q = team (first occurrence only)
    R = total goals (SUMIF, first occurrence only)
    S = rank (COUNTIFS-based, same pattern as Participant Standings col W)
    """
    ws.cell(row=4, column=4).value = (
        "Enter one row per goal scorer per match (multiple scorers per match OK)"
    )
    
    for row in range(6, 306):
        # P: player name only for first occurrence of this player in the log
        ws.cell(row=row, column=16).value = (
            f'=IF(OR(B{row}="",COUNTIF($B$6:$B{row},B{row})>1),"",B{row})'
        )
        # Q: team (only if P is populated)
        ws.cell(row=row, column=17).value = (
            f'=IF(P{row}="","",C{row})'
        )
        # R: total goals via SUMIF (only if P is populated)
        ws.cell(row=row, column=18).value = (
            f'=IF(P{row}="","",SUMIF($B$6:$B$305,B{row},$D$6:$D$305))'
        )
        # S: rank using COUNTIFS (same approach as Participant Standings)
        # Count first-occurrence rows with MORE goals + count same-goals rows up to here
        ws.cell(row=row, column=19).value = (
            f'=IF(P{row}="","",'
            f'COUNTIFS($R$6:$R$305,">"&R{row},$P$6:$P$305,"<>")'
            f'+COUNTIFS($R$6:$R{row},"="&R{row},$P$6:$P{row},"<>"))'
        )


# Fix .xlsx
print("Fixing .xlsx...")
wb = load_workbook('WorldCup2026_Sweepstake.xlsx')
fix_helper_columns(wb['Golden Boot'])
wb.save('WorldCup2026_Sweepstake.xlsx')
print("  Done")

# Fix .xlsm
print("Fixing .xlsm...")
try:
    wb2 = load_workbook('WorldCup2026_Sweepstake.xlsm', keep_vba=True)
    fix_helper_columns(wb2['Golden Boot'])
    wb2.save('WorldCup2026_Sweepstake.xlsm')
    print("  Done")
except PermissionError:
    print("  .xlsm is open in Excel - close it and re-run this script")

print("\nFormula sample (row 8):")
wb3 = load_workbook('WorldCup2026_Sweepstake.xlsx', data_only=False)
print(wb3['Golden Boot'].cell(row=8, column=19).value)
